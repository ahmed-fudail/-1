# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
from datetime import date, datetime
from pathlib import Path
from tkinter import Tk, filedialog

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "data" / "cost_centers_initial_data.json"
BACKUPS = ROOT / "data" / "backups"
IGNORED_SHEETS = {"تعليمات", "README", "ملخص"}

HEADERS = {
    "booking_number": ["رقم الحجز"],
    "cost_center": ["مركز التكلفة", "مركز التكلفه"],
    "amount_lyd": ['القيمة "د.ل"', "القيمة د.ل", "القيمه د.ل"],
    "contract_details": ["تفاصيل العقد"],
    "notes": ["ملاحظات"],
}


def normalize_digits(value: str) -> str:
    table: dict[int, str] = {}
    for digit in range(10):
        table[ord(chr(0x0660 + digit))] = str(digit)
        table[ord(chr(0x06F0 + digit))] = str(digit)
    return str(value).translate(table)


def normalize(value: object) -> str:
    text = normalize_digits(str(value or "").strip().lower())
    for old, new in [
        ("أ", "ا"), ("إ", "ا"), ("آ", "ا"), ("ى", "ي"),
        ("ؤ", "و"), ("ئ", "ي"), ("ة", "ه"),
    ]:
        text = text.replace(old, new)
    text = re.sub(r"[\u064B-\u065F\u0670\u0640]", "", text)
    return re.sub(r"[^\w]+", "", text, flags=re.UNICODE)


NORMALIZED_HEADERS = {
    field: {normalize(name) for name in names}
    for field, names in HEADERS.items()
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = normalize_digits(str(value)).strip()
    text = text.strip("\"'").strip()
    return "" if text.lower() in {"none", "null", "nan"} else text


def parse_amount(value: object) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_digits(str(value))
    for token in [",", "٬", "د.ل", "LYD"]:
        text = text.replace(token, "")
    try:
        return float(text.strip())
    except ValueError:
        return 0.0


def parse_note(cell, epoch) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, (int, float)) and cell.is_date:
        try:
            parsed = from_excel(value, epoch=epoch)
            if isinstance(parsed, (date, datetime)):
                return parsed.strftime("%d-%m-%Y")
        except Exception:
            pass
    text = clean_text(value)
    match = re.fullmatch(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", text)
    if match:
        day, month, year = map(int, match.groups())
        try:
            return date(year, month, day).strftime("%d-%m-%Y")
        except ValueError:
            pass
    return text


def choose_excel() -> Path:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askopenfilename(
        title="اختر ملف Excel لتحديث دليل مراكز التكلفة",
        filetypes=[("Excel Workbook", "*.xlsx")],
    )
    root.destroy()
    if not selected:
        raise SystemExit("تم إلغاء اختيار الملف.")
    return Path(selected)


def find_headers(ws) -> tuple[int, dict[str, int]]:
    for row in range(1, min(ws.max_row, 30) + 1):
        available = {
            normalize(ws.cell(row=row, column=col).value): col
            for col in range(1, ws.max_column + 1)
        }
        positions: dict[str, int] = {}
        for field, accepted in NORMALIZED_HEADERS.items():
            for name in accepted:
                if name in available:
                    positions[field] = available[name]
                    break
        if len(positions) == len(HEADERS):
            return row, positions
    raise ValueError("تعذر العثور على عناوين الأعمدة المطلوبة في ورقة نموذج تكلفة.")


def read_records(path: Path) -> tuple[list[dict], dict[str, int]]:
    workbook = load_workbook(path, data_only=True)
    records: list[dict] = []
    department_counts: dict[str, int] = {}
    seen_keys: set[tuple[str, str, str, str]] = set()

    for worksheet in workbook.worksheets:
        department = clean_text(worksheet.title)

        if not department or department in IGNORED_SHEETS:
            continue

        try:
            header_row, pos = find_headers(worksheet)
        except ValueError:
            print(f"تم تجاهل الصفحة «{department}» لعدم وجود العناوين المطلوبة.")
            continue

        department_count = 0

        for row in range(header_row + 1, worksheet.max_row + 1):
            booking = clean_text(worksheet.cell(row, pos["booking_number"]).value)
            center = clean_text(worksheet.cell(row, pos["cost_center"]).value)
            details = clean_text(worksheet.cell(row, pos["contract_details"]).value)
            amount = parse_amount(worksheet.cell(row, pos["amount_lyd"]).value)
            notes = parse_note(worksheet.cell(row, pos["notes"]), workbook.epoch)

            if not any([booking, center, details, amount, notes]):
                continue

            if not booking or not center or not details:
                print(
                    f"تحذير: تم تجاهل الصف {row} من قسم «{department}» "
                    "لوجود بيانات أساسية ناقصة."
                )
                continue

            duplicate_key = (
                normalize(department),
                normalize(booking),
                normalize(center),
                normalize(details),
            )

            if duplicate_key in seen_keys:
                print(
                    f"تحذير: تم تجاهل سجل مكرر في قسم «{department}» "
                    f"بالصف {row}: {booking}"
                )
                continue

            seen_keys.add(duplicate_key)
            department_count += 1

            records.append({
                "id": f"CC-{len(records) + 1:05d}",
                "department": department,
                "booking_number": booking,
                "cost_center": center,
                "amount_lyd": amount,
                "contract_details": details,
                "notes": notes,
                "source_sheet": department,
                "source_row": row,
                "is_archived": False,
                "needs_review": False,
            })

        department_counts[department] = department_count

    workbook.close()

    if not records:
        raise ValueError("لم يتم العثور على سجلات صالحة في صفحات الأقسام.")

    return records, department_counts


def backup_current() -> Path | None:
    if not OUTPUT.exists():
        return None
    BACKUPS.mkdir(parents=True, exist_ok=True)
    target = BACKUPS / f"cost_centers_{datetime.now():%Y%m%d_%H%M%S}.json"
    shutil.copy2(OUTPUT, target)
    return target


def write_json(records: list[dict], source: Path) -> None:
    payload = {
        "schema_version": 1,
        "dataset_name_ar": "دليل مراكز التكلفة",
        "source_file": source.name,
        "source_sheets": sorted({record["department"] for record in records}),
        "record_count": len(records),
        "currency": "LYD",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "records": records,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    temporary = OUTPUT.with_suffix(".json.tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    checked = json.loads(temporary.read_text(encoding="utf-8"))
    if checked["record_count"] != len(checked["records"]):
        raise ValueError("فشل التحقق من ملف JSON.")
    temporary.replace(OUTPUT)


def run(command: list[str], check: bool = True) -> subprocess.CompletedProcess:
    result = subprocess.run(
        command,
        cwd=ROOT,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
    )
    if result.stdout:
        print(result.stdout.strip())
    if result.stderr:
        print(result.stderr.strip())
    if check and result.returncode != 0:
        raise RuntimeError("فشل الأمر: " + " ".join(command))
    return result


def push_to_github(count: int, source_name: str) -> None:
    run(["git", "add", "data/cost_centers_initial_data.json"])
    diff = run(["git", "diff", "--cached", "--quiet"], check=False)
    if diff.returncode == 0:
        print("لا توجد تغييرات جديدة لرفعها.")
        return
    run(["git", "commit", "-m", f"Update Excel data ({count} records) from {source_name}"])
    run(["git", "push", "origin", "main"])
    print("تم رفع البيانات إلى GitHub بنجاح.")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    print("=" * 60)
    print("تحديث دليل مراكز التكلفة من Excel")
    print("=" * 60)

    excel_path = choose_excel()
    print(f"الملف المختار: {excel_path}")
    records, department_counts = read_records(excel_path)
    print("ملخص الأقسام:")
    for department, count in department_counts.items():
        print(f"  - {department}: {count} سجل")
    print(f"الإجمالي: {len(records)} سجلًا صالحًا.")

    backup = backup_current()
    if backup:
        print(f"تم إنشاء نسخة احتياطية: {backup.name}")

    write_json(records, excel_path)
    print(f"تم تحديث JSON: {OUTPUT}")

    if args.dry_run:
        print("اختبار ناجح: لم يتم تنفيذ Commit أو Push.")
    else:
        push_to_github(len(records), excel_path.name)

    print("اكتملت العملية بنجاح.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print("\nفشلت عملية التحديث")
        print(str(error))
        raise SystemExit(1)
